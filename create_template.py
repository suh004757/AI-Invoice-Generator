"""
Script to create sample Invoice template
"""
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from pathlib import Path


def create_sample_template():
    """Create a sample invoice template Excel file"""
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Invoice"
    
    # Set column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 18
    
    # Styles
    title_font = Font(name='Arial', size=20, bold=True)
    header_font = Font(name='Arial', size=11, bold=True)
    normal_font = Font(name='Arial', size=10)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    
    # Title
    ws['A1'] = 'INVOICE'
    ws['A1'].font = title_font
    ws.merge_cells('A1:E1')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30
    
    # Invoice Info Section
    ws['A3'] = 'Invoice No:'
    ws['B3'] = '[INVOICE_NO]'
    ws['A4'] = 'Date:'
    ws['B4'] = '[DATE]'
    ws['A5'] = 'Type:'
    ws['B5'] = '[TYPE]'
    
    # Customer Info Section
    ws['A7'] = 'Bill To:'
    ws['A7'].font = header_font
    ws['A8'] = 'Customer:'
    ws['B8'] = '[CUSTOMER_NAME]'
    ws['A9'] = 'Address:'
    ws['B9'] = '[CUSTOMER_ADDRESS]'
    
    # Items Table Header
    ws['A11'] = 'No.'
    ws['B11'] = 'Description'
    ws['C11'] = 'Quantity'
    ws['D11'] = 'Unit Price'
    ws['E11'] = 'Amount'
    
    for col in ['A', 'B', 'C', 'D', 'E']:
        cell = ws[f'{col}11']
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Sample item rows (12-19)
    for row in range(12, 20):
        for col in ['A', 'B', 'C', 'D', 'E']:
            cell = ws[f'{col}{row}']
            cell.border = border
            if col in ['C', 'D', 'E']:
                cell.alignment = Alignment(horizontal='right')
    
    # Totals Section
    ws['D21'] = 'Subtotal:'
    ws['D21'].font = header_font
    ws['E21'] = '[SUBTOTAL]'
    ws['E21'].alignment = Alignment(horizontal='right')
    
    ws['D22'] = 'VAT (10%):'
    ws['D22'].font = header_font
    ws['E22'] = '[VAT]'
    ws['E22'].alignment = Alignment(horizontal='right')
    
    ws['D23'] = 'Total:'
    ws['D23'].font = Font(name='Arial', size=12, bold=True)
    ws['E23'] = '[TOTAL]'
    ws['E23'].font = Font(name='Arial', size=12, bold=True)
    ws['E23'].alignment = Alignment(horizontal='right')
    ws['E23'].fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
    
    # Footer
    ws['A26'] = 'Payment Terms: Net 30 days'
    ws['A27'] = 'Thank you for your business!'
    ws['A27'].font = Font(name='Arial', size=10, italic=True)
    
    # Save template
    template_dir = Path(__file__).parent / "data"
    template_dir.mkdir(exist_ok=True)
    template_path = template_dir / "Invoice template.xlsx"
    
    wb.save(template_path)
    print(f"Template created: {template_path}")
    
    return template_path


if __name__ == "__main__":
    create_sample_template()
