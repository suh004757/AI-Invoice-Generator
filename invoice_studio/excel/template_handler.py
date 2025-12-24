"""
Excel template handler for Invoice Command Studio
"""
from openpyxl import load_workbook
from openpyxl.styles import Font
from pathlib import Path
from typing import Dict, List
from datetime import datetime
from .mapping import MappingConfig
from ..models.invoice import Invoice
from ..utils.formatters import format_currency, format_date, format_filename


class TemplateHandler:
    """Handle Excel template operations"""
    
    def __init__(self, template_path: str = None, mapping_config: MappingConfig = None):
        """
        Initialize template handler
        
        Args:
            template_path: Path to Excel template file
            mapping_config: Mapping configuration instance
        """
        if template_path is None:
            # Default to data/Invoice template.xlsx
            data_dir = Path(__file__).parent.parent.parent / "data"
            template_path = str(data_dir / "Invoice template.xlsx")
        
        self.template_path = template_path
        self.mapping = mapping_config or MappingConfig()
        
        # Verify template exists
        if not Path(self.template_path).exists():
            raise FileNotFoundError(f"Template not found: {self.template_path}")
    
    def generate_invoice(self, invoice: Invoice, output_dir: str = None) -> str:
        """
        Generate Excel invoice from template
        
        Args:
            invoice: Invoice object
            output_dir: Output directory (default: output/)
        
        Returns:
            Path to generated file
        """
        # Load template
        wb = load_workbook(self.template_path)
        ws = wb.active
        
        # Fill invoice data
        self._fill_header(ws, invoice)
        self._fill_items(ws, invoice)
        self._fill_totals(ws, invoice)
        
        # Apply VAT formatting based on type
        self._apply_vat_formatting(ws, invoice)
        
        # Generate filename
        if output_dir is None:
            output_dir = Path(__file__).parent.parent.parent / "output"
        
        output_dir = Path(output_dir)
        output_dir.mkdir(exist_ok=True)
        
        filename = format_filename(
            invoice.invoice_no,
            invoice.customer_name,
            invoice.type,
            invoice.date
        )
        
        output_path = output_dir / filename
        
        # Save file
        wb.save(output_path)
        
        return str(output_path)
    
    def _fill_header(self, ws, invoice: Invoice):
        """Fill invoice header information"""
        # Invoice number
        cell_ref = self.mapping.get('invoice_no', 'B3')
        ws[cell_ref] = invoice.invoice_no
        
        # Date
        cell_ref = self.mapping.get('date', 'B4')
        ws[cell_ref] = format_date(invoice.date)
        
        # Type
        cell_ref = self.mapping.get('type', 'B5')
        type_text = f"{invoice.type} Invoice"
        if invoice.type == "Tax":
            type_text += " (VAT 10%)"
        else:
            type_text += " (No VAT)"
        ws[cell_ref] = type_text
        
        # Customer name
        cell_ref = self.mapping.get('customer_name', 'B8')
        ws[cell_ref] = invoice.customer_name
        
        # Customer address (if available)
        cell_ref = self.mapping.get('customer_address', 'B9')
        ws[cell_ref] = ""  # Can be filled from customer master
    
    def _fill_items(self, ws, invoice: Invoice):
        """Fill invoice items table"""
        start_row = self.mapping.get('items_start_row', 12)
        cols = self.mapping.get('items_columns', {
            'no': 'A',
            'description': 'B',
            'quantity': 'C',
            'unit_price': 'D',
            'amount': 'E'
        })
        
        for idx, item in enumerate(invoice.items):
            row = start_row + idx
            
            # Item number
            ws[f"{cols['no']}{row}"] = idx + 1
            
            # Description
            ws[f"{cols['description']}{row}"] = item['description']
            
            # Quantity
            ws[f"{cols['quantity']}{row}"] = item['quantity']
            
            # Unit price
            ws[f"{cols['unit_price']}{row}"] = item['unit_price']
            
            # Amount
            ws[f"{cols['amount']}{row}"] = item['amount']
    
    def _fill_totals(self, ws, invoice: Invoice):
        """Fill totals section"""
        # Subtotal
        cell_ref = self.mapping.get('subtotal', 'E21')
        ws[cell_ref] = invoice.subtotal
        
        # VAT
        cell_ref = self.mapping.get('vat', 'E22')
        ws[cell_ref] = invoice.vat
        
        # Total
        cell_ref = self.mapping.get('total', 'E23')
        ws[cell_ref] = invoice.total
    
    def _apply_vat_formatting(self, ws, invoice: Invoice):
        """
        Apply VAT-specific formatting
        
        For Normal invoices (0% VAT), gray out or hide VAT row
        """
        vat_cell_ref = self.mapping.get('vat', 'E22')
        
        if invoice.type == "Normal":
            # Gray out VAT row
            # Get row number from cell reference
            import re
            match = re.search(r'\d+', vat_cell_ref)
            if match:
                vat_row = int(match.group())
                
                # Gray out the VAT label and value
                for col in ['D', 'E']:
                    cell = ws[f'{col}{vat_row}']
                    cell.font = Font(color='999999', italic=True)
    
    def load_template_for_preview(self):
        """
        Load template for preview/editing
        
        Returns:
            Workbook object
        """
        return load_workbook(self.template_path)
